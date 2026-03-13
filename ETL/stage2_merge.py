"""合併 STAGE1 CSV 檔案為單一 STAGE2 Excel 檔案."""

import argparse
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 標準化欄位名稱
STANDARD_COLUMNS = [
    "品牌",
    "產品分類", 
    "品名",
    "價格",
    "產品特色",
    "產品功效",
    "使用說明",
    "商品網址",
    "圖片網址",
    "爬蟲時間"
]

# 通用欄位對應（支援舊版/英文化欄位名稱）
BASE_COLUMN_MAPPING = {
    "品牌": "品牌",
    "brand": "品牌",
    "產品分類": "產品分類",
    "分類": "產品分類",
    "category": "產品分類",
    "品名": "品名",
    "name": "品名",
    "價格": "價格",
    "售價": "價格",
    "price": "價格",
    "說明文字": "產品特色",
    "商品描述": "產品特色",
    "產品特色": "產品特色",
    "description": "產品特色",
    "feature": "產品特色",
    "產品功效": "產品功效",
    "功效": "產品功效",
    "effect": "產品功效",
    "使用方式": "使用說明",
    "使用說明": "使用說明",
    "usage": "使用說明",
    "商品網址": "商品網址",
    "網址": "商品網址",
    "url": "商品網址",
    "圖片網址": "圖片網址",
    "主圖": "圖片網址",
    "image_url": "圖片網址",
    "爬蟲時間": "爬蟲時間",
    "爬取時間": "爬蟲時間",
    "crawled_at": "爬蟲時間",
}

# 特定品牌若有額外欄位名稱，可在此覆蓋
BRAND_COLUMN_OVERRIDES = {
    # "brand_key": {"自訂欄位": "標準欄位"}
}


def get_brand_from_filename(filename: str) -> str:
    """從檔案名稱取得品牌代碼."""
    name = Path(filename).stem
    return name.split("_stage")[0]


def standardize_dataframe(df: pd.DataFrame, brand_key: str) -> pd.DataFrame:
    """標準化 DataFrame 欄位."""
    normalized_brand = brand_key.lower()
    if normalized_brand.startswith("v_"):
        normalized_brand = normalized_brand[2:]  # 兼容 V_ 前綴的檔名
    
    # 先套用通用對應，再加入品牌專屬的覆蓋設定
    mapping = BASE_COLUMN_MAPPING.copy()
    mapping.update(BRAND_COLUMN_OVERRIDES.get(normalized_brand, {}))
    
    # 建立新的標準化 DataFrame
    standardized = pd.DataFrame()
    
    for std_col in STANDARD_COLUMNS:
        # 找到所有對應的原始欄位
        source_cols = [orig for orig, std in mapping.items() if std == std_col]
        valid_sources = [col for col in source_cols if col in df.columns]
        
        if valid_sources:
            # 若有多個欄位，合併內容（以空白分隔）
            if len(valid_sources) > 1:
                standardized[std_col] = df[valid_sources].fillna("").astype(str).apply(
                    lambda x: " ".join([s for s in x if s.strip()]), axis=1
                )
            else:
                standardized[std_col] = df[valid_sources[0]]
        else:
            standardized[std_col] = ""  # 填入空值
    
    return standardized


def merge_csvs(input_dir: Path, output_path: Path, stage: str = "stage1") -> None:
    """合併所有 STAGE1 CSV 檔案."""
    # 找出所有符合的 CSV 檔案
    csv_files = sorted(input_dir.glob(f"*_{stage}_*.csv"))
    
    if not csv_files:
        print(f"錯誤：在 {input_dir} 找不到任何 {stage} CSV 檔案")
        return
    
    all_data = []
    
    print("=" * 60)
    print(f"開始合併 {len(csv_files)} 個檔案...")
    print("=" * 60)
    
    for csv_file in csv_files:
        brand_key = get_brand_from_filename(csv_file.name)
        print(f"處理: {csv_file.name} (品牌: {brand_key})")
        
        try:
            df = pd.read_csv(csv_file, encoding="utf-8")
            standardized_df = standardize_dataframe(df, brand_key)
            all_data.append(standardized_df)
            print(f"  ✓ 成功讀取 {len(df)} 筆資料")
        except Exception as e:
            print(f"  ✗ 讀取失敗: {e}")
    
    if not all_data:
        print("錯誤：沒有成功讀取任何資料")
        return
    
    # 合併所有資料
    merged_df = pd.concat(all_data, ignore_index=True)
    
    # 寫入 Excel
    merged_df.to_excel(output_path, index=False, sheet_name="STAGE2_合併資料")
    
    # 格式化 Excel
    wb = load_workbook(output_path)
    ws = wb.active
    
    # 設定標題列樣式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 自動調整欄寬
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 凍結首列
    ws.freeze_panes = "A2"
    
    wb.save(output_path)
    
    print("=" * 60)
    print(f"合併完成！")
    print(f"總計: {len(merged_df)} 筆資料")
    print(f"輸出檔案: {output_path}")
    print("=" * 60)


def main():
    parser = argparse.ArgumentParser(description="合併 STAGE1 CSV 為 STAGE2 Excel")
    parser.add_argument(
        "--input-dir",
        type=str,
        default="output",
        help="輸入目錄（預設: output）"
    )
    parser.add_argument(
        "--stage",
        type=str,
        default="stage1",
        help="要合併的 stage（預設: stage1）"
    )
    args = parser.parse_args()
    
    # 取得路徑
    script_dir = Path(__file__).resolve().parent  # crawler/pipeline
    root_dir = script_dir.parent                 # crawler/

    input_dir = script_dir / args.input_dir
    if not input_dir.exists():
        alt_dir = root_dir / args.input_dir
        if alt_dir.exists():
            input_dir = alt_dir
            print(f"警告：{script_dir / args.input_dir} 不存在，改用 {input_dir}")
        else:
            print(f"錯誤：目錄 {input_dir} 或 {alt_dir} 不存在")
            return
    
    # 產生輸出檔名
    run_id = datetime.now().strftime("%y%m%d_%H%M")
    output_filename = f"merged_stage2_{run_id}.xlsx"
    output_path = input_dir / output_filename
    
    merge_csvs(input_dir, output_path, args.stage)


if __name__ == "__main__":
    main()

# 使用方式:
# python crawler/stage2_merge.py

